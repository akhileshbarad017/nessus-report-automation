#!/usr/bin/env python3
import argparse
import ipaddress
from pathlib import Path
import re
import sys
from collections import defaultdict

import pandas as pd
import pdfplumber
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter


ORDER = {"critical": 0, "high": 1, "medium": 2}
OUTPUT_COLUMNS = [
    "Vulnerability Name",
    "Risk Factor",
    "Affected IP:Port",
    "CVSS v3.0 Base Score",
    "CVE",
]

PLUGIN_RE = re.compile(r"^\s*(\d+)\s*-\s*(.+)$", re.M)
IP_RE = re.compile(r"\b(?:\d{1,3}\.){3}\d{1,3}\b")
PORT_RE = re.compile(r"(?:tcp|udp)/(\d{1,5})", re.I)
RISK_RE = re.compile(r"Risk Factor\s*\n\s*([A-Za-z]+)", re.I)
CVSS_RE = re.compile(r"CVSS v3\.0 Base Score\s*\n\s*([0-9]+(?:\.[0-9]+)?)", re.I)
CVE_RE = re.compile(r"CVE-\d{4}-\d{4,7}", re.I)

PLUGIN_OUTPUT_MARKER = "plugin output"
PLUGIN_OUTPUT_SCAN_LIMIT = 1500
HOST_SEARCH_WINDOW = 2000


def parse_args(argv):
    parser = argparse.ArgumentParser(
        description="Merge Nessus PDF findings into a single Excel report."
    )
    parser.add_argument(
        "--min-risk",
        choices=tuple(ORDER),
        default="medium",
        help="Only export findings at or above this risk level (default: medium).",
    )
    parser.add_argument(
        "--split-hosts",
        action="store_true",
        help="Export one row per host instead of a merged host list.",
    )
    parser.add_argument(
        "paths",
        nargs="+",
        help="Input PDF files followed by the output .xlsx filename.",
    )
    args = parser.parse_args(argv)

    if len(args.paths) < 2:
        parser.error("Provide at least one PDF input and one output .xlsx file.")

    excel_file = args.paths[-1]
    pdf_files = args.paths[:-1]

    if not excel_file.lower().endswith(".xlsx"):
        parser.error("Output must be a .xlsx file.")

    return pdf_files, excel_file, args.min_risk, args.split_hosts


def valid_port(port_text):
    try:
        port = int(port_text)
    except ValueError:
        return False
    return 1 <= port <= 65535


def find_host_before(text, pos, window=HOST_SEARCH_WINDOW):
    start = max(0, pos - window)
    snippet = text[start:pos]
    ips = IP_RE.findall(snippet)
    return ips[-1] if ips else None


def extract_words_text(page):
    try:
        words = page.extract_words()
    except Exception:
        return ""

    if not words:
        return ""

    ordered_words = sorted(words, key=lambda word: (round(word["top"], 1), word["x0"]))
    lines = []
    current_line = []
    current_top = None

    for word in ordered_words:
        word_top = round(word["top"], 1)
        if current_top is None or abs(word_top - current_top) <= 2:
            current_line.append(word["text"])
        else:
            lines.append(" ".join(current_line))
            current_line = [word["text"]]
        current_top = word_top

    if current_line:
        lines.append(" ".join(current_line))

    return "\n".join(lines)


def extract_page_text(page, stats):
    text = page.extract_text() or ""
    layout_text = page.extract_text(layout=True) or ""
    words_text = extract_words_text(page)

    candidates = [candidate for candidate in (text, layout_text, words_text) if candidate]
    if not candidates:
        return ""

    best_text = max(candidates, key=len)
    if best_text != text:
        stats["pages_fallback_used"] += 1
    return best_text


def extract_pdf_text(pdf_file, stats):
    with pdfplumber.open(pdf_file) as pdf:
        return "\n".join(extract_page_text(page, stats) for page in pdf.pages)


def extract_ports(block):
    marker_index = block.lower().find(PLUGIN_OUTPUT_MARKER)
    if marker_index == -1:
        return {"0"}

    search_block = block[marker_index : marker_index + PLUGIN_OUTPUT_SCAN_LIMIT]
    ports = {
        str(int(port_text))
        for port_text in PORT_RE.findall(search_block)
        if valid_port(port_text)
    }
    return ports or {"0"}


def init_stats():
    return {
        "plugins_found": 0,
        "plugins_exported": 0,
        "findings_exported": 0,
        "skipped_no_host": 0,
        "skipped_no_risk": 0,
        "skipped_unknown_risk": 0,
        "skipped_below_min_risk": 0,
        "warning_items": 0,
        "pages_fallback_used": 0,
    }


def build_accept_set(min_risk):
    threshold = ORDER[min_risk]
    return {risk for risk, rank in ORDER.items() if rank <= threshold}


def warning_path_for(excel_file):
    return str(Path(excel_file).with_suffix(".warnings.txt"))


def parse_host_port(host_port):
    host, sep, port = host_port.partition(":")
    port_value = int(port) if sep and port.isdigit() else -1

    try:
        ip_key = int(ipaddress.ip_address(host))
    except ValueError:
        ip_key = float("inf")

    return (ip_key, host, port_value, port)


def sorted_hosts(hosts):
    return sorted(hosts, key=parse_host_port)


def add_warning(warnings, stats, message):
    warnings.append(message)
    stats["warning_items"] += 1


def parse_findings(fulltext, vulns, accept, stats, warnings):
    matches = list(PLUGIN_RE.finditer(fulltext))
    total_plugins = len(matches)
    stats["plugins_found"] += total_plugins

    for idx, match in enumerate(matches):
        human_index = idx + 1
        if human_index % 20 == 0 or human_index == total_plugins:
            print(f"  Processing plugin {human_index}/{total_plugins}", end="\r")

        vuln_name = match.group(2).strip()
        block_start = match.end()
        block_end = matches[idx + 1].start() if idx + 1 < total_plugins else len(fulltext)
        block = fulltext[block_start:block_end]

        host_ip = find_host_before(fulltext, match.start())
        if not host_ip:
            stats["skipped_no_host"] += 1
            add_warning(warnings, stats, f"Missing host before plugin '{vuln_name}'")
            continue

        risk_match = RISK_RE.search(block)
        if not risk_match:
            stats["skipped_no_risk"] += 1
            add_warning(warnings, stats, f"Missing risk for plugin '{vuln_name}' on {host_ip}")
            continue

        risk = risk_match.group(1).strip().lower()
        if risk not in ORDER:
            stats["skipped_unknown_risk"] += 1
            add_warning(
                warnings,
                stats,
                f"Unrecognized risk '{risk}' for plugin '{vuln_name}' on {host_ip}",
            )
            continue

        if risk not in accept:
            stats["skipped_below_min_risk"] += 1
            continue

        cvss_match = CVSS_RE.search(block)
        cvss = cvss_match.group(1) if cvss_match else ""
        ports = extract_ports(block)

        for cve in CVE_RE.findall(block):
            vulns[vuln_name]["cve"].add(cve.upper())

        current = vulns[vuln_name]
        if not current["risk"] or ORDER[risk] < ORDER[current["risk"]]:
            current["risk"] = risk

        if cvss:
            try:
                if not current["cvss"] or float(cvss) > float(current["cvss"]):
                    current["cvss"] = cvss
            except ValueError:
                pass

        for port in ports:
            current["hosts"].add(f"{host_ip}:{port}")

        stats["plugins_exported"] += 1


def build_rows(vulns, split_hosts=False):
    rows = []
    for name, data in vulns.items():
        risk = data["risk"].lower()
        host_list = sorted_hosts(data["hosts"])

        if split_hosts:
            for host in host_list:
                rows.append(
                    {
                        "Vulnerability Name": name,
                        "Risk Factor": risk.title(),
                        "Affected IP:Port": host,
                        "CVSS v3.0 Base Score": data["cvss"],
                        "CVE": ", ".join(sorted(data["cve"])),
                    }
                )
            continue

        rows.append(
            {
                "Vulnerability Name": name,
                "Risk Factor": risk.title(),
                "Affected IP:Port": ", ".join(host_list),
                "CVSS v3.0 Base Score": data["cvss"],
                "CVE": ", ".join(sorted(data["cve"])),
            }
        )

    rows.sort(
        key=lambda row: (
            ORDER.get(row["Risk Factor"].lower(), len(ORDER)),
            row["Vulnerability Name"].lower(),
            parse_host_port(row["Affected IP:Port"].split(", ", 1)[0] or "0.0.0.0:0"),
        )
    )
    return rows


def write_excel(df, excel_file):
    with pd.ExcelWriter(excel_file, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Findings")
        worksheet = writer.sheets["Findings"]
        worksheet.freeze_panes = "A2"

        for cell in worksheet[1]:
            cell.font = Font(bold=True)

        for column_index, column_cells in enumerate(worksheet.iter_cols(), start=1):
            max_length = max(
                len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells
            )
            worksheet.column_dimensions[get_column_letter(column_index)].width = min(
                max_length + 2, 80
            )


def write_warnings(warnings, excel_file):
    if not warnings:
        return None

    warning_file = warning_path_for(excel_file)
    with open(warning_file, "w", encoding="utf-8") as handle:
        handle.write("\n".join(warnings))
        handle.write("\n")
    return warning_file


def main(argv=None):
    pdf_files, excel_file, min_risk, split_hosts = parse_args(argv or sys.argv[1:])
    vulns = defaultdict(lambda: {"risk": "", "cvss": "", "hosts": set(), "cve": set()})
    stats = init_stats()
    warnings = []
    accept = build_accept_set(min_risk)

    print("\nTotal PDFs:", len(pdf_files))
    print("Min risk:", min_risk.upper())
    print("Split hosts:", "ON" if split_hosts else "OFF")

    for pdf_index, pdf_file in enumerate(pdf_files, start=1):
        print(f"\n[{pdf_index}/{len(pdf_files)}] Reading:", pdf_file)
        try:
            fulltext = extract_pdf_text(pdf_file, stats)
        except FileNotFoundError:
            print(f"Error: file not found: {pdf_file}", file=sys.stderr)
            return 1
        except Exception as exc:
            print(f"Error reading {pdf_file}: {exc}", file=sys.stderr)
            return 1

        parse_findings(fulltext, vulns, accept, stats, warnings)

    print("\n")

    rows = build_rows(vulns, split_hosts=split_hosts)
    stats["findings_exported"] = len(rows)
    df = pd.DataFrame(rows, columns=OUTPUT_COLUMNS)

    print("Creating Excel file...")
    try:
        write_excel(df, excel_file)
    except Exception as exc:
        print(f"Error writing Excel file: {exc}", file=sys.stderr)
        return 1

    warning_file = write_warnings(warnings, excel_file)

    print("Done.")
    print("Merged from", len(pdf_files), "PDFs")
    print("File:", excel_file)
    print("Order -> CRITICAL -> HIGH -> MEDIUM")
    print("Plugins found:", stats["plugins_found"])
    print("Plugin entries exported:", stats["plugins_exported"])
    print("Exported findings rows:", stats["findings_exported"])
    print(
        "Skipped plugins:",
        stats["skipped_no_host"]
        + stats["skipped_no_risk"]
        + stats["skipped_unknown_risk"]
        + stats["skipped_below_min_risk"],
    )
    print("  Missing host:", stats["skipped_no_host"])
    print("  Missing risk:", stats["skipped_no_risk"])
    print("  Unrecognized risk:", stats["skipped_unknown_risk"])
    print("  Below min risk:", stats["skipped_below_min_risk"])
    print("Pages using fallback extraction:", stats["pages_fallback_used"])
    print("Warning items:", stats["warning_items"])
    if warning_file:
        print("Warnings file:", warning_file)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
